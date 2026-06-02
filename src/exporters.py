"""
exporters.py
------------
Exporta DataFrames y resultados a múltiples formatos:
Excel multi-hoja, CSV, JSON estructurado, GeoJSON.
"""

import json
import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
from typing import Dict, List, Any, Optional

HDR_FILL = PatternFill('solid', start_color='1F4E78')
HDR_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=11)
HDR_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
BODY_FONT = Font(name='Arial', size=10)
BORDER = Border(*[Side('thin', color='CCCCCC')]*4)
WRAP = Alignment(vertical='top', wrap_text=True)
CENTER = Alignment(horizontal='center', vertical='top', wrap_text=True)


def write_sheet(ws, df: pd.DataFrame, title: Optional[str] = None,
                subtitle: Optional[str] = None,
                conditional_cols: Optional[List[str]] = None,
                col_widths: Optional[Dict[str, int]] = None):
    """Escribe un DataFrame en una hoja con formato profesional."""
    row = 1
    if title:
        c = ws.cell(row=row, column=1, value=title)
        c.font = Font(name='Arial', bold=True, size=14, color='1F4E78')
        ws.merge_cells(start_row=row, start_column=1, end_row=row,
                       end_column=max(2, len(df.columns)))
        row += 1
    if subtitle:
        c = ws.cell(row=row, column=1, value=subtitle)
        c.font = Font(name='Arial', italic=True, size=10, color='666666')
        ws.merge_cells(start_row=row, start_column=1, end_row=row,
                       end_column=max(2, len(df.columns)))
        row += 1
    if title or subtitle:
        row += 1

    for c, col in enumerate(df.columns, 1):
        cell = ws.cell(row=row, column=c, value=str(col))
        cell.font = HDR_FONT; cell.fill = HDR_FILL
        cell.alignment = HDR_ALIGN; cell.border = BORDER
    hdr_row = row

    for r, (_, rrow) in enumerate(df.iterrows(), row+1):
        for c, col in enumerate(df.columns, 1):
            v = rrow[col]
            if pd.isna(v): v = ''
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = BODY_FONT; cell.border = BORDER
            cell.alignment = (CENTER if isinstance(v, (int, float))
                              and not isinstance(v, bool) else WRAP)

    last = row + len(df)
    if conditional_cols:
        for col_name in conditional_cols:
            if col_name in df.columns:
                cidx = list(df.columns).index(col_name) + 1
                letter = get_column_letter(cidx)
                rng = f'{letter}{hdr_row+1}:{letter}{last}'
                ws.conditional_formatting.add(rng, ColorScaleRule(
                    start_type='min', start_color='2E7D32',
                    mid_type='num', mid_value=0, mid_color='FFFFFF',
                    end_type='max', end_color='C62828',
                ))

    if col_widths:
        for col_name, width in col_widths.items():
            if col_name in df.columns:
                cidx = list(df.columns).index(col_name) + 1
                ws.column_dimensions[get_column_letter(cidx)].width = width


def to_excel(sheets: Dict[str, Dict[str, Any]], out_path: str):
    """Crea un Excel multi-hoja.

    sheets es un dict {nombre_hoja: {'df': df, 'title': str, 'subtitle': str,
                                      'conditional': [cols], 'widths': {col: w}}}
    """
    wb = Workbook()
    wb.remove(wb.active)
    for name, cfg in sheets.items():
        ws = wb.create_sheet(name[:31])  # Excel límite 31 chars
        write_sheet(
            ws, cfg['df'],
            title=cfg.get('title'),
            subtitle=cfg.get('subtitle'),
            conditional_cols=cfg.get('conditional'),
            col_widths=cfg.get('widths'),
        )
        ws.freeze_panes = 'A5' if cfg.get('title') else 'A2'
    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


def to_csv(df: pd.DataFrame, out_path: str):
    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    df.to_csv(out_path, index=False, encoding='utf-8')
    return out_path


def to_json_records(data: List[Dict], out_path: str):
    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    with open(out_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    return out_path


def to_geojson(places: List[Dict], out_path: str):
    """Crea un GeoJSON FeatureCollection desde una lista de lugares.

    Cada item debe tener: lat, lon, y propiedades adicionales.
    """
    features = []
    for p in places:
        lat, lon = p.get('lat'), p.get('lon')
        if lat is None or lon is None:
            continue
        props = {k: v for k, v in p.items() if k not in ('lat', 'lon')}
        features.append({
            'type': 'Feature',
            'geometry': {'type': 'Point', 'coordinates': [float(lon), float(lat)]},
            'properties': props,
        })
    geojson = {'type': 'FeatureCollection', 'features': features}
    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    with open(out_path, 'w', encoding='utf-8') as f:
        json.dump(geojson, f, ensure_ascii=False, indent=2)
    return out_path
